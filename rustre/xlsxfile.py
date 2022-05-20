#!/usr/bin/env python3
import os.path

from openpyxl import load_workbook
from openpyxl import Workbook


class XlsxFile:
    """Manage XSLX File operations.

        :param filename: xlsx filename to open (the file must exist)
        :type filename: str
        :param sheet_number: the zero based index of the xlsx sheet (default is 0)
        :type sheet_number: int
    """

    def __init__(self, filename, sheet_number=0):
        self.m_filename = filename
        if not os.path.exists(self.m_filename):
            raise ValueError("File: {} didn't exist!".format(self.m_filename))

        self.m_wb = load_workbook(self.m_filename)
        self.m_sheet = self.m_wb.worksheets[sheet_number]
        self.m_total_row = self.m_sheet.max_row
        self.m_total_col = self.m_sheet.max_column

    @staticmethod
    def create_file(filename):
        """Create an empty xlsx file

            :param filename: xlsx filename
            :type filename: str
            :return: Nothing
        """
        wb = Workbook()
        page = wb.active
        wb.save(filename)

    def get_columns(self, row_number=1):
        """Return all column values by row number

            :param row_number: the row index (start at 1)
            :type row_number: int
            :return: a list of all values in a row
            :rtype: list
        """
        return [cell.value for cell in self.m_sheet[row_number]]

    def get_row_count(self):
        """Return the total number of rows in the xlsx file

            :return: total number of rows
            :rtype: int
        """
        return self.m_total_row

    def append_row(self, row):
        """Append a row to the end of the xlsx file

            :param row: a list containing the values to append
            :type row: list
            :return: Nothing
        """
        self.m_sheet.append(row)

    def save(self):
        """Save any modification to the xlsx file

            :return: Nothing
        """
        self.m_wb.save(self.m_filename)

    def change_value(self, col, row, new_value):
        """Modify the value of a cell. The change is only saved after calling :method:save

            :param col: the column index (zero based)
            :type col: int
            :param row: the row index (1 based)
            :type row: int
            :param new_value:  the new cell value
            :type new_value: any
            :return: Nothing
        """
        self.m_sheet.cell(row, col).value = new_value
